"""
Excel Export Helper Module

Provides functionality to export pandas DataFrames to Excel files with formatting.
"""

from pathlib import Path
from typing import Optional, Union
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _get_peril_from_import_file(import_file: str) -> str:
    """
    Extract peril from Import File name.

    Peril is the first part before any underscore. For example:
    - "USEQ" -> "USEQ"
    - "USIF_Commercial" -> "USIF"
    - "USIF_Excess" -> "USIF"
    - "USFL_Other" -> "USFL"

    Args:
        import_file: Import File name

    Returns:
        Peril name (first part before underscore)
    """
    return import_file.split('_')[0]


def save_geohaz_validation_to_excel(
    validation_results: pd.DataFrame,
    date_value: str,
    cycle_type: str,
    output_dir: Union[str, Path]
) -> Optional[Path]:
    """
    Save GeoHaz validation results to an Excel file with one sheet per Peril.

    Creates a formatted Excel workbook where each sheet contains validation results
    for a single Peril (e.g., USEQ, USIF, USFL). Import Files are grouped by peril,
    which is the first part of the Import File name before any underscore.

    Includes formatting:
    - Bold headers
    - Conditional coloring on Status cell (green for PASS, red for FAIL)
    - Auto-fit column widths

    Args:
        validation_results: DataFrame from validate_geohaz_thresholds() with columns:
            Import File, Portfolio, Geocode Level, Expected %, Threshold %,
            Min %, Max %, Actual %, Risk Count, Status
        date_value: Date value for filename (e.g., '202511')
        cycle_type: Cycle type for filename (e.g., 'Quarterly')
        output_dir: Directory to save the Excel file

    Returns:
        Path to created Excel file, or None if validation_results is empty

    Example:
        ```python
        validation_results, all_passed = validate_geohaz_thresholds(
            geocoding_results, geohaz_thresholds, import_file_mapping
        )

        excel_path = save_geohaz_validation_to_excel(
            validation_results=validation_results,
            date_value='202511',
            cycle_type='Quarterly',
            output_dir=Path('/path/to/notebook/directory')
        )
        ```
    """
    # Don't create file if no results
    if validation_results.empty:
        return None

    # Ensure output_dir is a Path
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Build filename
    filename = f"GeoHaz_Validation_{date_value}_{cycle_type}.xlsx"
    file_path = output_dir / filename

    # Add Peril column based on Import File
    results_with_peril = validation_results.copy()
    results_with_peril['Peril'] = results_with_peril['Import File'].apply(_get_peril_from_import_file)

    # Get unique Perils
    perils = results_with_peril['Peril'].unique()

    # Create Excel writer
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for peril in perils:
            # Filter data for this Peril
            sheet_data = results_with_peril[
                results_with_peril['Peril'] == peril
            ].copy()

            # Drop the Peril column (redundant - it's the sheet name)
            sheet_data = sheet_data.drop(columns=['Peril'])

            # Write DataFrame to sheet
            sheet_data.to_excel(writer, sheet_name=peril, index=False)

            # Get the worksheet for formatting
            worksheet = writer.sheets[peril]

            # Apply formatting
            _format_validation_sheet(worksheet, sheet_data)

    return file_path


def _format_validation_sheet(worksheet, data: pd.DataFrame) -> None:
    """
    Apply formatting to a GeoHaz validation worksheet.

    Args:
        worksheet: openpyxl worksheet object
        data: DataFrame that was written to the sheet (for column info)
    """
    # Define styles
    header_font = Font(bold=True)
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    center_align = Alignment(horizontal='center')

    # Get column count and find Status column index
    num_cols = len(data.columns)
    status_col_idx = list(data.columns).index('Status') + 1  # 1-indexed for openpyxl

    # Format header row (row 1)
    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_align

    # Apply center alignment to data cells and conditional formatting to Status cell only
    for row in range(2, len(data) + 2):  # Start from row 2 (after header)
        # Center align all cells in the row
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = center_align

        # Apply conditional fill to Status cell only
        status_cell = worksheet.cell(row=row, column=status_col_idx)
        status_value = status_cell.value
        status_cell.fill = pass_fill if status_value == 'PASS' else fail_fill

    # Auto-fit column widths
    for col_idx, column in enumerate(data.columns, 1):
        # Calculate max width needed
        max_length = len(str(column))  # Header length
        for row in range(2, len(data) + 2):
            cell_value = worksheet.cell(row=row, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width with a little padding
        adjusted_width = max_length + 2
        worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def _is_flood_exposure_group(exposure_group: str) -> bool:
    """Check if an ExposureGroup is a Flood portfolio (USFL_*)."""
    return exposure_group.startswith('USFL_')


def save_control_totals_comparison_to_excel(
    comparison_results: pd.DataFrame,
    date_value: str,
    output_dir: Union[str, Path]
) -> Optional[Path]:
    """
    Save 3a vs 3b control totals comparison results to an Excel file.

    Creates a formatted Excel workbook with the comparison results showing
    the differences between Working Table (3a) and Contract Import File (3b)
    control totals. Results are split into two sheets:
    - 3a_vs_3b_NonFlood: Non-Flood perils (CBEQ, CBHU, USEQ, USFF, USST, USHU, USWF)
    - 3a_vs_3b_Flood: Flood perils (USFL_*)

    Includes formatting:
    - Bold headers
    - Conditional coloring on Status column (green for MATCH, red for MISMATCH)
    - Auto-fit column widths
    - Number formatting for difference columns

    Args:
        comparison_results: DataFrame from compare_3a_vs_3b_pivot() with columns:
            ExposureGroup, PolicyCount_Diff, PolicyPremium_Diff, PolicyLimit_Diff,
            LocationCountDistinct_Diff, TotalReplacementValue_Diff, LocationLimit_Diff,
            LocationDeductible_Diff, Status.
            Flood rows also have: AttachmentPoint_Diff, PolicyDeductible_Diff, PolicySublimit_Diff
        date_value: Date value for filename (e.g., '202503')
        output_dir: Directory to save the Excel file

    Returns:
        Path to created Excel file, or None if comparison_results is empty

    Example:
        ```python
        comparison_results, all_matched = compare_3a_vs_3b_pivot(
            results_3a, results_3b
        )

        excel_path = save_control_totals_comparison_to_excel(
            comparison_results=comparison_results,
            date_value='202503',
            output_dir=Path('/path/to/notebook/directory')
        )
        ```
    """
    # Don't create file if no results
    if comparison_results is None or comparison_results.empty:
        return None

    # Ensure output_dir is a Path
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Build filename
    filename = f"Control_Totals_Results_{date_value}.xlsx"
    file_path = output_dir / filename

    # Split results into Flood and Non-Flood
    is_flood = comparison_results['ExposureGroup'].apply(_is_flood_exposure_group)
    flood_results = comparison_results[is_flood].copy()
    non_flood_results = comparison_results[~is_flood].copy()

    # Define column order for each sheet
    # Non-Flood: 7 attributes + Status
    non_flood_columns = [
        'ExposureGroup',
        'PolicyCount_Diff',
        'PolicyPremium_Diff',
        'PolicyLimit_Diff',
        'LocationCountDistinct_Diff',
        'TotalReplacementValue_Diff',
        'LocationLimit_Diff',
        'LocationDeductible_Diff',
        'Status'
    ]

    # Flood: 10 attributes + Status (includes 3 Flood-specific)
    flood_columns = [
        'ExposureGroup',
        'PolicyCount_Diff',
        'PolicyPremium_Diff',
        'AttachmentPoint_Diff',
        'PolicyDeductible_Diff',
        'PolicyLimit_Diff',
        'PolicySublimit_Diff',
        'LocationCountDistinct_Diff',
        'TotalReplacementValue_Diff',
        'LocationLimit_Diff',
        'LocationDeductible_Diff',
        'Status'
    ]

    # Filter to only include columns that exist in each DataFrame
    non_flood_columns = [c for c in non_flood_columns if c in non_flood_results.columns]
    flood_columns = [c for c in flood_columns if c in flood_results.columns]

    # Reorder columns
    if not non_flood_results.empty:
        non_flood_results = non_flood_results[non_flood_columns]
    if not flood_results.empty:
        flood_results = flood_results[flood_columns]

    # Create Excel writer
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Write Non-Flood sheet
        if not non_flood_results.empty:
            non_flood_results.to_excel(writer, sheet_name='3a_vs_3b_NonFlood', index=False)
            worksheet = writer.sheets['3a_vs_3b_NonFlood']
            _format_comparison_sheet(worksheet, non_flood_results)

        # Write Flood sheet
        if not flood_results.empty:
            flood_results.to_excel(writer, sheet_name='3a_vs_3b_Flood', index=False)
            worksheet = writer.sheets['3a_vs_3b_Flood']
            _format_comparison_sheet(worksheet, flood_results)

    return file_path


def _format_comparison_sheet(worksheet, data: pd.DataFrame) -> None:
    """
    Apply formatting to a control totals comparison worksheet.

    Args:
        worksheet: openpyxl worksheet object
        data: DataFrame that was written to the sheet (for column info)
    """
    # Define styles
    header_font = Font(bold=True)
    match_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    mismatch_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')

    # Get column count
    num_cols = len(data.columns)

    # Find Status column index (if it exists)
    status_col_idx = None
    if 'Status' in data.columns:
        status_col_idx = list(data.columns).index('Status') + 1  # 1-indexed for openpyxl

    # Format header row (row 1)
    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_align

    # Apply formatting to data cells
    for row in range(2, len(data) + 2):  # Start from row 2 (after header)
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            col_name = data.columns[col - 1]

            # Right-align numeric columns (those ending with _Diff)
            if col_name.endswith('_Diff'):
                cell.alignment = right_align
            else:
                cell.alignment = center_align

        # Apply conditional fill to Status cell only
        if status_col_idx:
            status_cell = worksheet.cell(row=row, column=status_col_idx)
            status_value = status_cell.value
            if status_value == 'MATCH':
                status_cell.fill = match_fill
            elif status_value == 'MISMATCH':
                status_cell.fill = mismatch_fill

    # Auto-fit column widths
    for col_idx, column in enumerate(data.columns, 1):
        # Calculate max width needed
        max_length = len(str(column))  # Header length
        for row in range(2, len(data) + 2):
            cell_value = worksheet.cell(row=row, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width with a little padding
        adjusted_width = max_length + 2
        worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
