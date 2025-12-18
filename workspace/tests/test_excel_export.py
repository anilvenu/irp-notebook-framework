"""
Tests for excel_export module.
"""

import pandas as pd
from openpyxl import load_workbook
from helpers.excel_export import save_geohaz_validation_to_excel, _get_peril_from_import_file


class TestGetPerilFromImportFile:
    """Tests for _get_peril_from_import_file function."""

    def test_peril_extraction(self):
        """Test peril extraction from Import File names."""
        # No underscore - returns as-is
        assert _get_peril_from_import_file('USEQ') == 'USEQ'
        assert _get_peril_from_import_file('CBHU') == 'CBHU'

        # With underscore - returns first part
        assert _get_peril_from_import_file('USIF_Commercial') == 'USIF'
        assert _get_peril_from_import_file('USIF_Excess') == 'USIF'
        assert _get_peril_from_import_file('USFL_Other') == 'USFL'
        assert _get_peril_from_import_file('USHU_Full_Extra') == 'USHU'


class TestSaveGeohazValidationToExcel:
    """Tests for save_geohaz_validation_to_excel function."""

    def test_groups_import_files_by_peril(self, tmp_path):
        """Test that Import Files are grouped by Peril into same sheet."""
        validation_results = pd.DataFrame([
            {
                'Import File': 'USIF_Commercial',
                'Portfolio': 'USIF_Commercial',
                'Geocode Level': 'Coordinate',
                'Expected %': 93.0,
                'Threshold %': 5.0,
                'Min %': 88.0,
                'Max %': 98.0,
                'Actual %': 93.2,
                'Risk Count': 932,
                'Status': 'PASS'
            },
            {
                'Import File': 'USIF_Excess',
                'Portfolio': 'USIF_Excess',
                'Geocode Level': 'Coordinate',
                'Expected %': 90.0,
                'Threshold %': 5.0,
                'Min %': 85.0,
                'Max %': 95.0,
                'Actual %': 89.0,
                'Risk Count': 890,
                'Status': 'PASS'
            },
            {
                'Import File': 'USEQ',
                'Portfolio': 'USEQ',
                'Geocode Level': 'Coordinate',
                'Expected %': 95.0,
                'Threshold %': 3.0,
                'Min %': 92.0,
                'Max %': 98.0,
                'Actual %': 94.0,
                'Risk Count': 940,
                'Status': 'PASS'
            }
        ])

        result = save_geohaz_validation_to_excel(
            validation_results=validation_results,
            date_value='202511',
            cycle_type='Quarterly',
            output_dir=tmp_path
        )

        # Verify file created with correct name
        assert result is not None
        assert result.exists()
        assert result.name == 'GeoHaz_Validation_202511_Quarterly.xlsx'

        # Verify sheets grouped by peril (2 sheets, not 3)
        wb = load_workbook(result)
        assert len(wb.sheetnames) == 2
        assert 'USIF' in wb.sheetnames
        assert 'USEQ' in wb.sheetnames

        # Verify USIF sheet has both USIF_Commercial and USIF_Excess
        ws_usif = wb['USIF']
        assert ws_usif.cell(row=2, column=1).value == 'USIF_Commercial'
        assert ws_usif.cell(row=3, column=1).value == 'USIF_Excess'

        wb.close()

    def test_returns_none_for_empty_results(self, tmp_path):
        """Test that empty validation results returns None and creates no file."""
        validation_results = pd.DataFrame(columns=[
            'Import File', 'Portfolio', 'Geocode Level', 'Expected %',
            'Threshold %', 'Min %', 'Max %', 'Actual %', 'Risk Count', 'Status'
        ])

        result = save_geohaz_validation_to_excel(
            validation_results=validation_results,
            date_value='202511',
            cycle_type='Quarterly',
            output_dir=tmp_path
        )

        assert result is None
        assert not list(tmp_path.glob('*.xlsx'))
